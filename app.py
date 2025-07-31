from flask import Flask, render_template, request
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from datetime import datetime

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xlsx', 'csv', 'zip'}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        email = request.form.get('email')
        files = request.files.getlist('file')

        if not email or not files:
            return "Inserisci email e almeno un file.", 400

        safe_email = secure_filename(email.replace("@", "_at_"))
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_email)
        os.makedirs(save_path, exist_ok=True)

        log_file = 'upload_log.xlsx'
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Crea il file di log se non esiste
        if not os.path.exists(log_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Email ente", "Data invio", "Nome file"])
            wb.save(log_file)
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)

        # Carica il file di log
        wb = load_workbook(log_file)
        ws = wb.active

        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(save_path, filename)
                file.save(filepath)

                # Scrivi nel log
                ws.append([email, now, filename])

        wb.save(log_file)

        return render_template('upload_success.html', email=email)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
